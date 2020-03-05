from PyQt5 import QtWidgets, uic, QtCore, QtGui

from PyQt5.QtWidgets import QFileDialog, QMessageBox
from translate.storage.tmx import tmxfile
from translate.storage.xliff import xlifffile, xliffunit
from lxml import etree
import sys
import openpyxl
import os
import re

delimiter = '|'
def display_error_message(content):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setInformativeText(content)
    msg.setWindowTitle("Error")
    msg.exec_()

def get_file_type(extension):
    if extension == ".txt":
        extension = "Text"
    if extension == ".xlsx":
        extension = "Excel"
    if extension == ".xliff":
        extension = "Xliff"
    if extension == ".tmx":
        extension = "Tmx"
    return extension


CONVERT_TYPE = ["Text to Excel", "Excel to Text", "Text to Tmx", "Tmx to Text", "Text to Xliff", "Xliff to Text", "Excel to Xliff", "Xliff to Excell", \
    "Excel to Tmx", "Tmx to Excel", "Xliff to Tmx", "Tmx to Xliff"]

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        uic.loadUi("Main.ui", self)
        self.selected_task_id = -1
        self.tbl_task.setHorizontalHeaderLabels(['Source File', 'Destination File', 'Convert Type'])
        header = self.tbl_task.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)

    @QtCore.pyqtSlot()
    def create_task(self):
        dialog = Dialog()
        if dialog.exec_():# accept, save item
            rows = self.tbl_task.rowCount()
            self.tbl_task.insertRow(rows)
            source_item = QtWidgets.QTableWidgetItem(dialog.txt_source.text())
            source_item.setToolTip(dialog.txt_source.text())
            self.tbl_task.setItem(rows, 0, source_item)
            dest_item = QtWidgets.QTableWidgetItem(dialog.txt_destination.text())
            dest_item.setToolTip(dialog.txt_destination.text())
            self.tbl_task.setItem(rows, 1, dest_item)
            convert_type = ""
            filename, file_extension = os.path.splitext(dialog.txt_source.text())
            
            convert_type += get_file_type(file_extension) + " to "
            filename, file_extension = os.path.splitext(dialog.txt_destination.text())
            convert_type += get_file_type(file_extension)            
            self.tbl_task.setItem(rows, 2, QtWidgets.QTableWidgetItem(convert_type))
        else:# reject
            pass
        
    @QtCore.pyqtSlot()
    def remove_task(self):
        if self.selected_task_id != -1:
            self.tbl_task.removeRow(self.selected_task_id)
            self.selected_task_id = -1
    @QtCore.pyqtSlot()
    def run_task(self):
        if self.selected_task_id != -1:
            source_file = self.tbl_task.item(self.selected_task_id, 0).text()
            dest_file = self.tbl_task.item(self.selected_task_id, 1).text()
            c_type = self.tbl_task.item(self.selected_task_id, 2).text()
            if c_type == "Text to Excel":
                self.convert_Text2Excel(source_file, dest_file)
            if c_type == "Excel to Text":
                self.convert_Excel2Text(source_file, dest_file)
            if c_type == "Text to Tmx":
                self.convert_Text2Tmx(source_file, dest_file)
            if c_type == "Tmx to Text":
                self.convert_Tmx2Text(source_file, dest_file)
            if c_type == "Text to Xliff":
                self.convert_Text2Xliff(source_file, dest_file)
            if c_type == "Xliff to Text":
                self.convert_Xliff2Text(source_file, dest_file)
            if c_type == "Excel to Tmx":
                self.convert_Excel2Tmx(source_file, dest_file)
            if c_type == "Tmx to Excel":
                self.convert_Tmx2Excel(source_file, dest_file)
            if c_type == "Tmx to Xliff":
                self.convert_Tmx2Xliff(source_file, dest_file)
            if c_type == "Xliff to Tmx":
                self.convert_Xliff2Tmx(source_file, dest_file)
            if c_type == "Excel to Xliff":
                self.convert_Excel2Xliff(source_file, dest_file)
            if c_type == "Xliff to Excel":
                self.convert_Xliff2Excel(source_file, dest_file)


    @QtCore.pyqtSlot()
    def select_item(self):
        self.selected_task_id = self.tbl_task.currentRow()
    def convert_Text2Excel(self, src, dst):
        if not os.path.exists(dst):
            dst_wb = openpyxl.Workbook()
            ss_sheet = dst_wb['Sheet']
            ss_sheet.title = 'transmem'
            dst_wb.save(dst)
        dst_wb = openpyxl.load_workbook(dst)
        dst_ws = dst_wb['transmem']
        lines = open(src, encoding='utf-8').read().strip().split('\n')
        dst_ws.cell(1, 1).value = 'en'
        dst_ws.cell(1, 2).value = 'th'
        
        for row, l in enumerate(lines, start=2):
            s = l.split(delimiter)
            dst_ws.cell(row, 1).value = s[0]
            dst_ws.cell(row, 2).value = s[1]
        dst_wb.save(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Excel2Text(self, src, dst):
        src_wb = openpyxl.load_workbook(src)
        src_ws = src_wb.worksheets[0]
        
        en_col = 1
        th_col = 2
        for col in range(1,src_ws.max_column):
            cell_value = src_ws.cell(1, col).value.lower()
            if 'en' == cell_value:
                en_col = col
            if 'th' == cell_value:
                th_col = col
        with open(dst, 'w', encoding='utf-8') as dst_file:
            for row in range(2, src_ws.max_row+1):
                dst_file.write(src_ws.cell(row, en_col).value + delimiter + src_ws.cell(row, th_col).value + '\n')
        dst_file.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")
    
    def convert_Text2Tmx(self, src, dst):
        tmx_file = tmxfile()
        lines = open(src, encoding='utf-8').read().strip().split('\n')
        for line in lines:
            s = line.split(delimiter)
            tmx_file.addtranslation(s[0], "en", s[1], "th")
        tmx_file.savefile(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Tmx2Text(self, src, dst):
        dst_file = open(dst, 'w', encoding='utf-8')
        with open(src, 'rb') as fin:
            tmx_file = tmxfile(fin, 'en', 'th')
            for node in tmx_file.unit_iter():
                dst_file.write(node.getsource() + delimiter + node.gettarget() + '\n')
        dst_file.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Text2Xliff(self, src, dst):
        xliff_file = xlifffile()
        xliff_file.setsourcelanguage('en')
        xliff_file.settargetlanguage('th')
        lines = open(src, encoding='utf-8').read().strip().split('\n')
        for line in lines:
            s = line.split(delimiter)
            node = xliffunit(s[0])
            node.settarget(s[1])
            xliff_file.addunit(node)
        
        xliff_file.savefile(dst)
        fin = open(dst, "r", encoding='utf-8')
        data = fin.read()
        fin.close()
        data = data.replace('<xliff xmlns="urn:oasis:names:tc:xliff:document:1.1" version="1.1">', '<xliff xmlns="urn:oasis:names:tc:xliff:document:1.2" version="1.2">')
        fout = open(dst, 'w', encoding='utf-8')
        fout.write(data)
        fout.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")
    def convert_Xliff2Text(self, src, dst):
        fin = open(src, 'r', encoding = "utf-8")
        data = fin.read()
        xliff_file = xlifffile.parsestring(data)
        txt_file = open(dst, 'w', encoding='utf-8')
        for node in xliff_file.unit_iter():
            txt_file.write(node.source + delimiter + node.target + '\n')
        txt_file.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Excel2Tmx(self, src, dst):
        src_wb = openpyxl.load_workbook(src)
        src_ws = src_wb.worksheets[0]
        
        en_col = 1
        th_col = 2
        for col in range(1,src_ws.max_column):
            cell_value = src_ws.cell(1, col).value.lower()
            if 'en' == cell_value:
                en_col = col
            if 'th' == cell_value:
                th_col = col
        tmx_file = tmxfile()
        for row in range(2, src_ws.max_row+1):
            tmx_file.addtranslation(src_ws.cell(row, en_col).value, "en", src_ws.cell(row, th_col).value, "th")
        tmx_file.savefile(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")
    
    def convert_Tmx2Excel(self, src, dst):
        if not os.path.exists(dst):
            dst_wb = openpyxl.Workbook()
            ss_sheet = dst_wb['Sheet']
            ss_sheet.title = 'transmem'
            dst_wb.save(dst)
        dst_wb = openpyxl.load_workbook(dst)
        dst_ws = dst_wb['transmem']
        lines = open(src, encoding='utf-8').read().strip().split('\n')
        dst_ws.cell(1, 1).value = 'en'
        dst_ws.cell(1, 2).value = 'th'
        with open(src, 'rb') as fin:
            tmx_file = tmxfile(fin, 'en', 'th')
            row = 2
            for node in tmx_file.unit_iter():
                dst_ws.cell(row, 1).value = node.getsource()
                dst_ws.cell(row, 2).value = node.gettarget()
                row += 1
        dst_wb.save(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")
    def convert_Tmx2Xliff(self, src, dst):
        xliff_file = xlifffile()
        xliff_file.setsourcelanguage('en')
        xliff_file.settargetlanguage('th')
        with open(src, 'rb') as fin:
            tmx_file = tmxfile(fin, 'en', 'th')
            for node in tmx_file.unit_iter():
                new_node = xliffunit(node.getsource())
                new_node.settarget(node.gettarget())
                xliff_file.addunit(new_node)
        xliff_file.savefile(dst)
        fin = open(dst, "r", encoding='utf-8')
        data = fin.read()
        fin.close()
        data = data.replace('<xliff xmlns="urn:oasis:names:tc:xliff:document:1.1" version="1.1">', '<xliff xmlns="urn:oasis:names:tc:xliff:document:1.2" version="1.2">')
        fout = open(dst, 'w', encoding='utf-8')
        fout.write(data)
        fout.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")
    def convert_Xliff2Tmx(self, src, dst):
        fin = open(src, 'r', encoding = "utf-8")
        data = fin.read()
        xliff_file = xlifffile.parsestring(data)
        tmx_file = tmxfile()
        for node in xliff_file.unit_iter():
            tmx_file.addtranslation(node.source, "en", node.target, "th")
        tmx_file.savefile(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Excel2Xliff(self, src, dst):
        src_wb = openpyxl.load_workbook(src)
        src_ws = src_wb.worksheets[0]        
        en_col = 1
        th_col = 2
        for col in range(1,src_ws.max_column):
            cell_value = src_ws.cell(1, col).value.lower()
            if 'en' == cell_value:
                en_col = col
            if 'th' == cell_value:
                th_col = col
        xliff_file = xlifffile()
        xliff_file.setsourcelanguage('en')
        xliff_file.settargetlanguage('th')
        for row in range(2, src_ws.max_row+1):
            new_node = xliffunit(src_ws.cell(row, en_col).value)
            new_node.settarget(src_ws.cell(row, th_col).value)
            xliff_file.addunit(new_node)
        xliff_file.savefile(dst)
        fin = open(dst, "r", encoding='utf-8')
        data = fin.read()
        fin.close()
        data = data.replace('<xliff xmlns="urn:oasis:names:tc:xliff:document:1.1" version="1.1">', '<xliff xmlns="urn:oasis:names:tc:xliff:document:1.2" version="1.2">')
        fout = open(dst, 'w', encoding='utf-8')
        fout.write(data)
        fout.close()
        QMessageBox.information(self, "Information", "Converting was done successfully")

    def convert_Xliff2Excel(self, src, dst):
        if not os.path.exists(dst):
            dst_wb = openpyxl.Workbook()
            ss_sheet = dst_wb['Sheet']
            ss_sheet.title = 'transmem'
            dst_wb.save(dst)
        dst_wb = openpyxl.load_workbook(dst)
        dst_ws = dst_wb['transmem']
        lines = open(src, encoding='utf-8').read().strip().split('\n')
        dst_ws.cell(1, 1).value = 'en'
        dst_ws.cell(1, 2).value = 'th'
        fin = open(src, 'r', encoding = "utf-8")
        data = fin.read()
        xliff_file = xlifffile.parsestring(data)
        row = 2
        for node in xliff_file.unit_iter():
            dst_ws.cell(row, 1).value = node.source
            dst_ws.cell(row, 2).value = node.target
            row += 1
        dst_wb.save(dst)
        QMessageBox.information(self, "Information", "Converting was done successfully")

class Dialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(Dialog, self).__init__(parent)
        uic.loadUi("Task.ui", self)
    @QtCore.pyqtSlot()
    def browse_source(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Select Source File", "","Available File(*.txt *.xlsx *.tmx *.xliff);;Bilingual Text Files (*.txt);;Excel File(*.xlsx);;Tmx File(*.tmx);;Xliff file(*.xliff)", options=options)
        if fileName:
            self.txt_source.setText(fileName)

    @QtCore.pyqtSlot()
    def browse_destination(self):
        extension_list = ['.txt', '.xlsx', '.tmx', '.xliff']
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, extension = QFileDialog.getSaveFileName(self,"Select Destination File","","Bilingual Text Files (*.txt);;Excel File(*.xlsx);;Tmx File(*.tmx);;Xliff file(*.xliff)", options=options)        
        if fileName:
            for ext in extension_list:
                if ext in extension and not fileName.endswith(ext):
                    fileName += ext
                    break
            self.txt_destination.setText(fileName)
    @QtCore.pyqtSlot()
    def save(self):
        if self.txt_source.text() == "":
            display_error_message("Please put the source file")
        elif self.txt_destination.text() == "":
            display_error_message("Please put the destination file")
        else:
            source_path = self.txt_source.text()
            dest_path = self.txt_destination.text()
            s_filename, s_file_extension = os.path.splitext(self.txt_source.text())
            d_filename, d_file_extension = os.path.splitext(self.txt_destination.text())
            if source_path == dest_path:
                QMessageBox.warning(self, "Warning", "File paths are the same")
                return
            if s_file_extension == d_file_extension:
                QMessageBox.warning(self, "Warning", "Extensions are the same")
                return
            self.accept()

if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication([])
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())

